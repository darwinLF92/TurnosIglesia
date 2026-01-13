# views.py

from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import CreateView, UpdateView, ListView
from django.urls import reverse_lazy
from .models import Procesion
from .forms import ProcesionForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models import Count, Q
from django.http import JsonResponse
from gestion_turnos.models import RegistroInscripcion
from turnos.models import Turno
from django.db.models.functions import ExtractYear
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
import tempfile
import openpyxl
from django.views.decorators.http import require_POST
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter



@method_decorator(login_required, name='dispatch')
class ListaProcesionesView(ListView):
    model = Procesion
    template_name = 'procesiones/lista_procesiones.html'
    context_object_name = 'procesiones'

    def get_queryset(self):
        return Procesion.objects.filter(activo=True).annotate(
            total_turnos=Count('turnos', filter=Q(turnos__activo=True))
        ).order_by('-fecha_creacion')   # 👈 más reciente primero

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 👇 AQUÍ se valida el grupo (esto SÍ funciona)
        context['es_admin'] = self.request.user.groups.filter(
            name='Administrador'
        ).exists()

        return context



@method_decorator(login_required, name='dispatch')
class CrearProcesionView(CreateView):
    model = Procesion
    form_class = ProcesionForm
    template_name = 'procesiones/crear_procesion.html'

    def form_valid(self, form):
        """
        Si el formulario es válido, guarda la procesión y muestra un mensaje de éxito.
        """
        form.save()
        messages.success(self.request, "Procesión creada correctamente.")
        return self.render_to_response(self.get_context_data(form=form, success=True, message="Procesión creada correctamente."))

    def form_invalid(self, form):
        """
        Si hay errores en el formulario, los muestra en SweetAlert sin redirigir.
        """
        messages.error(self.request, "Hubo un error al crear la procesión. Inténtelo de nuevo.")
        return self.render_to_response(self.get_context_data(form=form, success=False))

@method_decorator(login_required, name='dispatch')
class EditarProcesionView(UpdateView):
    model = Procesion
    form_class = ProcesionForm
    template_name = 'procesiones/editar_procesion.html'

    def form_valid(self, form):
        """
        Si el formulario es válido, actualiza la procesión y muestra un mensaje de éxito.
        """
        form.save()
        messages.success(self.request, "Procesión actualizada correctamente.")
        return self.render_to_response(self.get_context_data(form=form, success=True, message="Procesión actualizada correctamente."))

    def form_invalid(self, form):
        """
        Si hay errores en el formulario, los muestra en SweetAlert sin redirigir.
        """
        messages.error(self.request, "Hubo un error al actualizar la procesión. Inténtelo de nuevo.")
        return self.render_to_response(self.get_context_data(form=form, success=False))

@login_required
def EliminarProcesionView(request, pk):
    procesion = get_object_or_404(Procesion, pk=pk)
    if request.method == 'POST':
        procesion.activo = False
        procesion.save()
        return redirect('procesiones:lista_procesiones')
    return render(request, 'procesiones/eliminar_procesion.html', {'procesion': procesion})


def reporte_turnos(request):
    anio = request.GET.get('anio')
    procesion_id = request.GET.get('procesion_id')
    turno_id = request.GET.get('turno_id')

    if not procesion_id:
        return JsonResponse({'error': 'Procesión no seleccionada'}, status=400)

    procesion = Procesion.objects.filter(id=procesion_id).first()
    if not procesion:
        return JsonResponse({'error': 'Procesión no encontrada'}, status=404)

    turnos = Turno.objects.filter(procesion_id=procesion_id).order_by('numero_turno')
    if turno_id:
        turnos = turnos.filter(id=turno_id)

    data = []
    for turno in turnos:
        inscritos = RegistroInscripcion.objects.filter(turno=turno).count()
        entregados = RegistroInscripcion.objects.filter(turno=turno, entregado=True).count()
        data.append({
            'numero_turno': turno.numero_turno,
            'referencia': turno.referencia,
            'marcha_funebre': turno.marcha_funebre,
            'inscritos': inscritos,
            'entregados': entregados,
        })

    return JsonResponse({
        'procesion_nombre': procesion.nombre,
        'fecha_procesion': procesion.fecha.strftime('%d/%m/%Y'),
        'turnos': data
    })

def obtener_anios(request):
    anios = Procesion.objects.annotate(anio=ExtractYear('fecha')) \
        .values_list('anio', flat=True).distinct().order_by('-anio')
    return JsonResponse(list(anios), safe=False)

def exportar_reporte_turnos_pdf(request):
    anio = request.GET.get('anio')
    procesion_id = request.GET.get('procesion_id')
    turno_id = request.GET.get('turno_id')

    if not procesion_id:
        return HttpResponse("Procesión no seleccionada", status=400)

    procesion = Procesion.objects.filter(id=procesion_id).first()
    if not procesion:
        return HttpResponse("Procesión no encontrada", status=404)

    turnos = Turno.objects.filter(procesion_id=procesion_id).order_by('numero_turno')
    if turno_id:
        turnos = turnos.filter(id=turno_id)

    datos = []
    for turno in turnos:
        inscritos = RegistroInscripcion.objects.filter(turno=turno).count()
        entregados = RegistroInscripcion.objects.filter(turno=turno, entregado=True).count()
        datos.append({
            'numero_turno': turno.numero_turno,
            'referencia': turno.referencia,
            'marcha_funebre': turno.marcha_funebre,
            'inscritos': inscritos,
            'entregados': entregados,
        })

    html_string = render_to_string('procesiones/reporte_turnos_pdf.html', {
        'procesion': procesion,
        'anio': anio,
        'turnos': datos
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_turnos.pdf"'

    with tempfile.NamedTemporaryFile(delete=True) as tmp_file:
        HTML(string=html_string).write_pdf(target=response)

    return response

def exportar_reporte_turnos_excel(request): 
    anio = request.GET.get('anio')
    procesion_id = request.GET.get('procesion_id')
    turno_id = request.GET.get('turno_id')

    if not procesion_id:
        return HttpResponse("Procesión no seleccionada", status=400)

    procesion = Procesion.objects.filter(id=procesion_id).first()
    if not procesion:
        return HttpResponse("Procesión no encontrada", status=404)

    turnos = Turno.objects.filter(procesion_id=procesion_id).order_by('numero_turno')

    if turno_id:
        turnos = turnos.filter(id=turno_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Turnos"

    # === Estilos básicos ===
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="1F4E78")  # azul oscuro
    header_font = Font(bold=True, color="FFFFFF")         # blanco
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # === Encabezado de la procesión ===
    ws["A1"] = "Nombre Procesión:"
    ws["B1"] = procesion.nombre
    ws["A2"] = "Fecha Procesión:"
    ws["B2"] = procesion.fecha.strftime("%d/%m/%Y")

    ws["A1"].font = ws["A2"].font = bold_font

    # Un pequeño título encima (opcional)
    ws.insert_rows(1)
    ws["A1"] = "REPORTE DE TURNOS"
    ws.merge_cells("A1:J1")
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # La fila de encabezados quedará en la fila 5
    ws.append([])  # fila 3 vacía
    ws.append([])  # fila 4 vacía

    ws.append([
        "Turno No.",
        "Tipo de turno",
        "Clase de turno",
        "Reservado",
        "Reservado para",
        "Referencia",
        "Marcha Fúnebre",
        "Inscritos",
        "Entregados",
        "Costo del turno",
    ])
    header_row = ws.max_row

    # Aplicar estilo al encabezado
    for col in range(1, 11):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    total_costo = 0

    # === Filas de datos ===
    for turno in turnos:
        inscritos = RegistroInscripcion.objects.filter(turno=turno).count()
        entregados = RegistroInscripcion.objects.filter(turno=turno, entregado=True).count()

        # Reservado / Reservado para
        if turno.reservado_hermandad:
            reservado_str = "Sí"
            nombre_limpio = (turno.nombre_hermandad_visitante or "").strip()
            if nombre_limpio:
                reservado_para = nombre_limpio
            else:
                reservado_para = "Extraordinario"
        else:
            reservado_str = "No"
            reservado_para = ""

        costo_turno = turno.valor
        total_costo += turno.valor * inscritos


        ws.append([
            turno.numero_turno,
            turno.get_tipo_turno_display(),
            turno.get_clase_turno_display(),
            reservado_str,
            reservado_para,
            turno.referencia or "",
            turno.marcha_funebre or "",
            inscritos,
            entregados,
            float(costo_turno),
        ])

        # Borde a la fila recién agregada
        current_row = ws.max_row
        for col in range(1, 11):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border

    # === Fila de TOTAL ===
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=9, value="TOTAL:")
    ws.cell(row=total_row, column=9).font = bold_font
    ws.cell(row=total_row, column=10, value=float(total_costo))
    ws.cell(row=total_row, column=10).font = bold_font
    ws.cell(row=total_row, column=10).border = thin_border

    # === Formato moneda para la columna de costo ===
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=10)
        cell.number_format = '"Q"#,##0.00'

    # === Ajustar ancho de columnas automáticamente ===
    for col in range(1, 11):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_turnos.xlsx"'
    wb.save(response)
    return response



@login_required
@require_POST
def marcar_procesion_relevante(request):
    if not request.user.groups.filter(name='Administrador').exists():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    procesion_id = request.POST.get('procesion_id')
    procesion = get_object_or_404(Procesion, id=procesion_id)

    # Si se marca como relevante, desmarcar las demás
    if not procesion.es_relevante:
        Procesion.objects.exclude(id=procesion.id).update(es_relevante=False)
        procesion.es_relevante = True
    else:
        procesion.es_relevante = False

    procesion.save()

    return JsonResponse({
        'success': True,
        'es_relevante': procesion.es_relevante,
        'procesion_id': procesion.id
    })
