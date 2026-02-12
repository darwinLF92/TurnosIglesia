from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from .forms import InscripcionForm
from .models import RegistroInscripcion, Turno
from django.views.generic import ListView
from django.http import JsonResponse
from devotos.models import Devoto
from procesiones.models import Procesion
from django.utils.dateparse import parse_date
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.views import View
from weasyprint import HTML
import tempfile
from django.template.loader import get_template
from establecimiento.models import Establecimiento
import os
import urllib.parse  # Importar para codificar URLs
from django.conf import settings
from django.contrib import messages
from django.utils import timezone
import json
from weasyprint import CSS
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models.functions import ExtractYear
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import io
from django.db import transaction
from nucleo.correos import enviar_confirmacion_inscripcion_correo


@login_required
def inscribir_devoto(request):
    can_turno_reservado = request.user.has_perm('establecimiento.turno_reservado')

    devotos_activos = Devoto.objects.filter(activo=True).order_by('nombre')
    procesiones = Procesion.objects.filter(activo=True).order_by('-fecha_creacion')

    if request.method == 'POST':
        form = InscripcionForm(request.POST, user=request.user)

        if not form.is_valid():
            # 👇 IMPORTANTE: aquí no redirigimos, mostramos errores
# Captura errores reales del form (incluye __all__)
            errores = form.errors.get_json_data()

            # Construir un texto amigable con todos los mensajes
            lista = []
            for campo, errores_campo in errores.items():
                for e in errores_campo:
                    # campo == '__all__' -> error general
                    if campo == '__all__':
                        lista.append(e.get("message"))
                    else:
                        lista.append(f"{campo}: {e.get('message')}")

            mensaje = "<br>".join([m for m in lista if m]) or "No se pudo registrar la inscripción. Verifique los datos."

            messages.error(request, mensaje)

            return render(request, 'gestion_turnos/crear_inscripcion.html', {
                'form': form,
                'devotos_activos': devotos_activos,
                'procesiones': procesiones,
                'can_turno_reservado': can_turno_reservado,
            })


        try:
            with transaction.atomic():
                inscripcion = form.save(commit=False)

                # 🔒 Turno reservado sin permiso -> NO redirigir, solo mostrar error
                if (
                    inscripcion.turno
                    and getattr(inscripcion.turno, "reservado_hermandad", False)
                    and not can_turno_reservado
                ):
                    messages.error(request, "No tiene permiso para inscribir en turnos reservados.")
                    return render(request, 'gestion_turnos/crear_inscripcion.html', {
                        'form': form,
                        'devotos_activos': devotos_activos,
                        'procesiones': procesiones,
                        'can_turno_reservado': can_turno_reservado,
                    })

                # ✅ Tomar valores ya validados del form
                fecha_estimada = form.cleaned_data.get("fecha_entrega_estimada")
                lugar_entrega = (form.cleaned_data.get("lugar_entrega") or "").strip()

                if fecha_estimada and settings.USE_TZ and timezone.is_naive(fecha_estimada):
                    fecha_estimada = timezone.make_aware(fecha_estimada, timezone.get_current_timezone())

                # ✅ Guardar SIEMPRE en inscripción
                inscripcion.fecha_entrega_estimada = fecha_estimada
                inscripcion.lugar_entrega = lugar_entrega or None

                # ✅ Guardar en Turno solo si está vacío y el usuario ingresó
                if inscripcion.turno:
                    turno_obj = inscripcion.turno
                    actualizar_turno = False

                    if not turno_obj.fecha_entrega and fecha_estimada:
                        turno_obj.fecha_entrega = fecha_estimada
                        actualizar_turno = True

                    if not turno_obj.lugar_entrega and lugar_entrega:
                        turno_obj.lugar_entrega = lugar_entrega
                        actualizar_turno = True

                    if actualizar_turno:
                        turno_obj.save(update_fields=["fecha_entrega", "lugar_entrega"])

                # Control
                inscripcion.inscrito = True
                inscripcion.tipo_inscripcion = "presencial"

                if inscripcion.turno:
                    inscripcion.valor_turno = inscripcion.turno.valor

                    procesion = inscripcion.turno.procesion
                    max_turnos_local = getattr(procesion, "turnos_devoto_local", 0) or 0

                    if max_turnos_local > 0:
                        inscripciones_devoto_en_procesion = RegistroInscripcion.objects.filter(
                            devoto=inscripcion.devoto,
                            turno__procesion=procesion,
                            inscrito=True,
                        ).count()

                        if inscripciones_devoto_en_procesion >= max_turnos_local:
                            messages.error(
                                request,
                                f"No se puede inscribir a un devoto a más {max_turnos_local} turnos, contactar al Administrador para sugerir o solicitar cambio."
                            )
                            return render(request, 'gestion_turnos/crear_inscripcion.html', {
                                'form': form,
                                'devotos_activos': devotos_activos,
                                'procesiones': procesiones,
                                'can_turno_reservado': can_turno_reservado,
                            })

                inscripcion.cambio = inscripcion.calcular_cambio()
                inscripcion.save()

            messages.success(request, "Inscripción registrada correctamente.")
            return redirect(reverse('gestion_turnos:lista_inscripciones'))

        except Exception as e:
      
            messages.error(request, f"Ocurrió un error al registrar la inscripción: {str(e)}")
            return render(request, 'gestion_turnos/crear_inscripcion.html', {
                'form': form,
                'devotos_activos': devotos_activos,
                'procesiones': procesiones,
                'can_turno_reservado': can_turno_reservado,
            })

    # GET
    form = InscripcionForm(user=request.user)
    return render(request, 'gestion_turnos/crear_inscripcion.html', {
        'form': form,
        'devotos_activos': devotos_activos,
        'procesiones': procesiones,
        'can_turno_reservado': can_turno_reservado,
    })


@login_required
def obtener_precio_turno(request):
    turno_id = request.GET.get('turno_id')

    if not turno_id:
        return JsonResponse({"error": "turno_id requerido"}, status=400)

    try:
        turno = Turno.objects.get(id=turno_id, activo=True)
    except Turno.DoesNotExist:
        return JsonResponse({"error": "Turno no encontrado"}, status=404)

    fecha = turno.fecha_entrega  # DateTimeField

    # ✅ Formato para <input type="datetime-local"> -> "YYYY-MM-DDTHH:MM"
    if fecha:
        if timezone.is_aware(fecha):
            fecha = timezone.localtime(fecha)
        # si es naive, la dejamos tal cual
        fecha_str = fecha.strftime("%Y-%m-%dT%H:%M")
    else:
        fecha_str = ""

    return JsonResponse({
        "precio": str(turno.valor),
        "fecha_entrega": fecha_str,
        "lugar_entrega": turno.lugar_entrega or "",
        "reservado": bool(turno.reservado_hermandad),
    })

@method_decorator(login_required, name='dispatch')
class ListaInscripciones(ListView): 
    model = RegistroInscripcion
    template_name = "gestion_turnos/lista_inscripciones.html"
    context_object_name = "inscripciones"
    paginate_by = 7

    def get_queryset(self):
        # Mostrar solo inscripciones activas (inscrito=True)
        queryset = RegistroInscripcion.objects.filter(
            inscrito=True
        ).select_related("devoto", "turno__procesion").order_by("-fecha_inscripcion")

        # Capturar los parámetros de búsqueda desde la URL
        nombre = self.request.GET.get("nombre", "").strip()
        cui_o_nit = self.request.GET.get("cui_o_nit", "").strip()
        fecha_inicio = self.request.GET.get("fecha_inicio", "").strip()
        fecha_fin = self.request.GET.get("fecha_fin", "").strip()

        if nombre:
            queryset = queryset.filter(devoto__nombre__icontains=nombre)

        if cui_o_nit:
            queryset = queryset.filter(devoto__cui_o_nit__icontains=cui_o_nit)

        if fecha_inicio:
            fecha_inicio = parse_date(fecha_inicio)
            if fecha_inicio:
                queryset = queryset.filter(fecha_inscripcion__gte=fecha_inicio)

        if fecha_fin:
            fecha_fin = parse_date(fecha_fin)
            if fecha_fin:
                queryset = queryset.filter(fecha_inscripcion__lte=fecha_fin)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        page = self.request.GET.get("page", 1)
        paginator = Paginator(self.get_queryset(), self.paginate_by)
        page_obj = paginator.get_page(page)

        context["inscripciones"] = page_obj
        return context


@login_required
def load_turnos(request):
    procesion_id = request.GET.get('procesion_id')

    # 🔐 Permiso para turnos reservados
    can_turno_reservado = request.user.has_perm('establecimiento.turno_reservado')

    turnos = Turno.objects.filter(
        procesion_id=procesion_id,
        activo=True
    ).order_by('numero_turno')

    data = []
    for turno in turnos:
        es_reservado = bool(turno.reservado_hermandad)  # ✅ tu campo real

        data.append({
            "id": turno.id,
            "numero_turno": turno.numero_turno,
            "referencia": turno.referencia or "",
            "valor": str(turno.valor),
            "reservado": es_reservado,
            "habilitado": (not es_reservado) or can_turno_reservado,

            # 👇 DEBUG (quita luego)
            "debug_perm": can_turno_reservado,
            "debug_reservado": es_reservado,
            "tipo_turno": (turno.tipo_turno),
        })

    return JsonResponse(data, safe=False)

@login_required
def ajax_devotos_activos(request):
    devotos = Devoto.objects.filter(activo=True).values('id', 'nombre', 'cui_o_nit')
    return JsonResponse(list(devotos), safe=False)

@login_required
def ajax_procesiones_activas(request):
    procesiones = Procesion.objects.filter(activo=True).values('id', 'nombre')
    return JsonResponse(list(procesiones), safe=False)

@login_required
def buscar_devotos_list(request):
    termino_busqueda = request.GET.get('q', '')
    devotos = Devoto.objects.filter(nombre__icontains=termino_busqueda).values('nombre')[:10]  # Limita los resultados a 10
    return JsonResponse(list(devotos), safe=False)

# Vista para anular una inscripción
@method_decorator(login_required, name='dispatch')
class AnularInscripcionView(View):
    def get(self, request, inscripcion_id):
        """Renderiza la plantilla de confirmación de anulación."""
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)
        return render(request, 'gestion_turnos/anular_inscripcion.html', {'inscripcion': inscripcion})

    def post(self, request, inscripcion_id):
        """Procesa la anulación de la inscripción."""
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)

        if not inscripcion.inscrito:
            messages.error(request, "Esta inscripción ya ha sido anulada.")
            return render(request, 'gestion_turnos/anular_inscripcion.html', {'inscripcion': inscripcion})

        # Obtener el turno asociado
        turno = inscripcion.turno

        # **Actualizar los valores del turno**
        if turno:
            turno_inscritos_actualizados = RegistroInscripcion.objects.filter(turno=turno, inscrito=True).count() - 1
            turno_disponibles_actualizados = turno.capacidad - turno_inscritos_actualizados

            # **Evitar números negativos**
            turno_inscritos_actualizados = max(turno_inscritos_actualizados, 0)
            turno_disponibles_actualizados = max(turno_disponibles_actualizados, 0)

            # **Guardar cambios en el turno**
            turno.save()

        # **Anular inscripción**
        inscripcion.inscrito = False
        inscripcion.entregado = False
        inscripcion.fecha_entrega = None
        inscripcion.save()

        messages.success(request, "Inscripción anulada correctamente.")
        return render(request, 'gestion_turnos/anular_inscripcion.html', {'inscripcion': inscripcion})
# Vista para generar el comprobante de inscripción
@method_decorator(login_required, name='dispatch')
class ComprobanteInscripcionView(View):
    def get(self, request, inscripcion_id):
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)
        html = render_to_string('gestion_turnos/comprobante_inscripcion.html', {'inscripcion': inscripcion})
        return HttpResponse(html)
    
@method_decorator(login_required, name='dispatch')
class ComprobanteRecepcionView(View):
    def get(self, request, inscripcion_id):
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)
        html = render_to_string('gestion_turnos/comprobante_recepcion.html', {'inscripcion': inscripcion})
        return HttpResponse(html)


# Vista para el detalle de inscripción
@method_decorator(login_required, name='dispatch')
class DetalleInscripcionView(View):
    def get(self, request, inscripcion_id):
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)
        return render(request, 'gestion_turnos/detalle_inscripcion.html', {'inscripcion': inscripcion})
    
@method_decorator(login_required, name='dispatch')
class ComprobanteInscripcionPDFView(View):
    def get(self, request, inscripcion_id):
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)
        establecimiento = Establecimiento.objects.filter(estado='activo').first()

        html_string = render_to_string(
            "gestion_turnos/comprobante_inscripcion_pdf.html",
            {'inscripcion': inscripcion, 'establecimiento': establecimiento}
        )

        pdf_content = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="comprobante_{inscripcion.id}.pdf"'
        return response

@method_decorator(login_required, name='dispatch')    
class ComprobanteRecepcionPDFView(View):
    def get(self, request, inscripcion_id):
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)
        establecimiento = Establecimiento.objects.filter(estado='activo').first()

        html_string = render_to_string(
            "gestion_turnos/comprobante_recepcion_pdf.html",
            {'inscripcion': inscripcion, 'establecimiento': establecimiento}
        )

        pdf_content = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="recepcion_{inscripcion.id}.pdf"'
        return response


@method_decorator(login_required, name='dispatch')
class EnviarComprobanteWhatsAppView(View):
    def get(self, request, inscripcion_id):
        # Obtener la inscripción
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)

        # Obtener el teléfono del devoto
        telefono_devoto = inscripcion.devoto.telefono
        if not telefono_devoto:
            return JsonResponse({"error": "El devoto no tiene un número de teléfono registrado."}, status=400)

        # Limpiar el número (solo dígitos)
        telefono_devoto = ''.join(filter(str.isdigit, telefono_devoto))

        # Obtener la procesión y turno
        nombre_procesion = inscripcion.turno.procesion.nombre
        numero_turno = inscripcion.turno.numero_turno
        marcha_funebre= inscripcion.turno.marcha_funebre
        referencia= inscripcion.turno.referencia
        fecha_procesion= inscripcion.turno.procesion.fecha.strftime("%d-%m-%Y")
        fecha_entrega_estimada = inscripcion.fecha_entrega_estimada.strftime("%d-%m-%Y") if inscripcion.fecha_entrega_estimada else "No definida"
        hora_entrega = inscripcion.fecha_entrega_estimada.strftime("%H:%M") if inscripcion.fecha_entrega_estimada else ""
        lugar_entrega = inscripcion.lugar_entrega if inscripcion.lugar_entrega else "No definido"

        # Obtener el establecimiento
        establecimiento = Establecimiento.objects.filter(estado='activo').first()
        nombre_hermandad = establecimiento.hermandad if establecimiento else "la Hermandad"

        # Generar el PDF y guardarlo en MEDIA_ROOT
        pdf_filename = f"comprobante_{inscripcion_id}.pdf"
        pdf_path = os.path.join(settings.MEDIA_ROOT, pdf_filename)
        
        HTML(string=render_to_string("gestion_turnos/comprobante_inscripcion_pdf.html", {'inscripcion': inscripcion})).write_pdf(target=pdf_path)

        # Obtener la URL del archivo PDF y asegurar que use HTTPS
        pdf_url = request.build_absolute_uri(settings.MEDIA_URL + pdf_filename)
        pdf_url = pdf_url.replace("http://", "http://")

        # Construir el mensaje de WhatsApp con los datos relevantes
        mensaje = (
            f"Hola {inscripcion.devoto.nombre},\n\n"
            f"✅ Se ha inscrito exitosamente a la procesión *{nombre_procesion}* de fecha: *{fecha_procesion }*, en el turno *{numero_turno}* - *{referencia}*.\n\n"
            f"🎶 *Marcha Fúnebre:* {marcha_funebre}\n"
            f"📅 *Fecha de entrega:* {fecha_entrega_estimada} a las {hora_entrega} Horas.\n"
            f"🔐 *Contraseña del Turno:* {inscripcion.codigo}\n"
            f"📍 *Lugar de entrega:* {lugar_entrega}\n\n"
            f"Agradecemos su colaboración con nuestra Hermandad y, de parte de Nuestro Señor Jesucristo, le deseamos bendiciones.\n\n"
            f"🙏 Que tenga un buen y bendecido Turno.\n\n"
            f"Atentamente,\n{nombre_hermandad}.\n\n"
        )

        mensaje_codificado = urllib.parse.quote(mensaje)
        whatsapp_url = f"https://wa.me/{telefono_devoto}?text={mensaje_codificado}"


        # Retornar la URL de WhatsApp
        return JsonResponse({
            "whatsapp_url": whatsapp_url,
            "mensaje": mensaje  # ← Mensaje sin urlencode
        }, status=200)

    
@method_decorator(login_required, name='dispatch')
class ListaEntregaTurnos(View):
    template_name = "gestion_turnos/lista_entrega_turnos.html"

    def get(self, request):
        entregados = request.GET.get("entregados", "").lower() == "true"
        nombre = request.GET.get("nombre", "").strip()

        # Filtrar inscripciones por entregados y nombre de devoto
        inscripciones = (
            RegistroInscripcion.objects.filter(inscrito=True, entregado=entregados)
            .select_related("devoto", "turno")
            .order_by("-fecha_inscripcion")  # ✅ más reciente -> más antiguo
        )

        if nombre:
            inscripciones = inscripciones.filter(devoto__nombre__icontains=nombre)

        # Paginación
        paginacion = request.GET.get("paginacion", 6)  # Se permite personalizar cantidad de elementos por página
        paginator = Paginator(inscripciones, paginacion)
        page = request.GET.get("page")
        inscripciones_paginadas = paginator.get_page(page)

        # Mantener los filtros en la paginación
        query_params = f"&nombre={nombre}&entregados={entregados}"

        return render(request, self.template_name, {
            "inscripciones": inscripciones_paginadas,
            "filtro_entregados": entregados,
            "nombre": nombre,
            "query_params": query_params,  # Se pasa para mantener los filtros en la paginación
        })



@method_decorator(login_required, name='dispatch')
class EntregarTurnoView(View):
    def post(self, request, inscripcion_id):
        inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id, inscrito=True, entregado=False)

        # Marcar como entregado
        inscripcion.entregado = True
        inscripcion.fecha_entrega = timezone.now()
        inscripcion.save()

        messages.success(request, f"Turno entregado correctamente a {inscripcion.devoto.nombre}.")
        return JsonResponse({"success": True})

@method_decorator(login_required, name='dispatch')   
class ValidarEntregaTurnoView(View):
    def post(self, request, inscripcion_id):
        try:
            data = json.loads(request.body)
            codigo_ingresado = data.get("id_inscripcion")

            # Buscar la inscripción válida (activa, no entregada)
            inscripcion = get_object_or_404(
                RegistroInscripcion,
                id=inscripcion_id,
                inscrito=True,
                entregado=False
            )

            # Validar usando el campo 'codigo'
            if str(inscripcion.codigo).strip() != str(codigo_ingresado).strip():
                return JsonResponse({
                    "success": False,
                    "message": "El código ingresado no coincide con la inscripción. Verifique e intente nuevamente."
                }, status=400)

            # Marcar como entregado
            inscripcion.entregado = True
            inscripcion.fecha_entrega = timezone.now()
            inscripcion.save()

            return JsonResponse({
                "success": True,
                "message": "Turno entregado correctamente."
            })

        except json.JSONDecodeError:
            return JsonResponse({
                "success": False,
                "message": "Error en la solicitud. Intente nuevamente."
            }, status=400)
        
from datetime import datetime
from django.db.models import Count, Sum, Q


@login_required
def reporte_inscripciones(request):
    # Filtros
    filtro_procesion = request.GET.get('procesion', '')
    filtro_turno = request.GET.get('turno')
    fecha_inicio_str = request.GET.get('fechainicio', '')
    fecha_fin_str = request.GET.get('fechafin', '')
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')

    # Validar fechas solo si se enviaron
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else None
    except ValueError:
        fecha_inicio = None

    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None
    except ValueError:
        fecha_fin = None

    # Obtener años únicos de procesiones
    anios = (Procesion.objects
        .filter(activo=True)
        .annotate(anio=ExtractYear('fecha'))
        .values_list('anio', flat=True)
        .distinct()
        .order_by('-anio')
    )

    # Inscripciones activas
    inscripciones = RegistroInscripcion.objects.filter(
    inscrito=True,
    turno__procesion__activo=True
    )

    if fecha_inicio:
        inscripciones = inscripciones.filter(fecha_inscripcion__date__gte=fecha_inicio)
    if fecha_fin:
        inscripciones = inscripciones.filter(fecha_inscripcion__date__lte=fecha_fin)

    if filtro_procesion:
        inscripciones = inscripciones.filter(turno__procesion_id=filtro_procesion)

    if filtro_turno:
        inscripciones = inscripciones.filter(turno_id=filtro_turno)

    # Agrupación por procesión
    procesiones = inscripciones.values(
        'turno__procesion__id',
        'turno__procesion__nombre'
    ).annotate(
        total_turnos=Count('turno', distinct=True),
        total_inscritos=Count('id'),
        total_entregados=Count('id', filter=Q(entregado=True)),
        total_valor=Sum('valor_turno'),
        total_pagado=Sum('monto_pagado')
    ).order_by('turno__procesion__nombre')

    # Detalle por turno
    turnos = inscripciones.values(
        'turno__id',
        'turno__numero_turno',
        'turno__procesion__nombre'
    ).annotate(
        cantidad_inscritos=Count('id'),
        entregados=Count('id', filter=Q(entregado=True)),
        total_valor=Sum('valor_turno'),
        total_pagado=Sum('monto_pagado')
    ).order_by('turno__procesion__nombre', 'turno__numero_turno')

    # Lista de devotos por turno
    devotos_por_turno = inscripciones.values(
        'turno__id',
        'turno__numero_turno',
        'devoto__id',
        'devoto__nombre',
        'fecha_inscripcion',
        'entregado',
        'valor_turno',
        'monto_pagado',
        'cambio'
    ).order_by('turno__id', 'devoto__nombre')

    # Lista general de devotos inscritos
    devotos_inscritos = inscripciones.values(
        'devoto__id',
        'devoto__nombre',
        'turno__numero_turno',
        'turno__procesion__nombre',
        'fecha_inscripcion',
        'entregado',
        'monto_pagado',
        'cambio'
    ).order_by('devoto__nombre')

    return JsonResponse({
        'fecha_hoy': fecha_hoy,
        'procesiones': list(procesiones),
        'turnos': list(turnos),
        'anios': list(anios),
        'devotos_por_turno': list(devotos_por_turno),
        'devotos_inscritos': list(devotos_inscritos),
    })



def obtener_anios_procesiones(request):
    anios = (Procesion.objects
        .filter(activo=True)
        .annotate(anio=ExtractYear('fecha'))
        .values_list('anio', flat=True)
        .distinct()
        .order_by('-anio')
    )
    return JsonResponse(list(anios), safe=False)

def obtener_procesiones_por_anio(request):
    anio = request.GET.get('anio')

    if not anio or not anio.isdigit():
        return JsonResponse({'error': 'Año inválido.'}, status=400)

    procesiones = (Procesion.objects
        .filter(fecha__year=int(anio), activo=True)
        .values('id', 'nombre')
        .order_by('nombre')
    )
    return JsonResponse(list(procesiones), safe=False)


def obtener_turnos_por_procesion(request):
    """
    Devuelve los turnos asociados a una procesión.
    """
    procesion_id = request.GET.get('procesion_id')

    if not procesion_id or not procesion_id.isdigit():
        return JsonResponse({'error': 'ID de procesión inválido.'}, status=400)

    turnos = Turno.objects.filter(procesion_id=int(procesion_id), activo=True).values('id', 'numero_turno')
    return JsonResponse(list(turnos), safe=False)

@login_required
def exportar_inscripciones_pdf(request):
    filtro_procesion = request.GET.get('procesion', '')
    filtro_turno = request.GET.get('turno')
    fecha_inicio_str = request.GET.get('fechainicio', '')
    fecha_fin_str = request.GET.get('fechafin', '')
    fecha_hoy = datetime.now()

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else None
    except ValueError:
        fecha_inicio = None

    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None
    except ValueError:
        fecha_fin = None

    inscripciones = RegistroInscripcion.objects.filter(
        inscrito=True,
        turno__procesion__activo=True
    )

    if fecha_inicio:
        inscripciones = inscripciones.filter(fecha_inscripcion__date__gte=fecha_inicio)
    if fecha_fin:
        inscripciones = inscripciones.filter(fecha_inscripcion__date__lte=fecha_fin)
    if filtro_procesion:
        inscripciones = inscripciones.filter(turno__procesion_id=filtro_procesion)
    if filtro_turno:
        inscripciones = inscripciones.filter(turno_id=filtro_turno)

    devotos = inscripciones.order_by('turno__numero_turno', 'devoto__nombre')

    # Contar inscritos
    if filtro_turno:
        cantidad_inscritos = RegistroInscripcion.objects.filter(
            turno_id=filtro_turno,
            inscrito=True,
            turno__procesion__activo=True
        ).count()
    elif filtro_procesion:
        turnos = Turno.objects.filter(procesion_id=filtro_procesion, activo=True)  # si Turno tiene activo
        cantidad_inscritos = RegistroInscripcion.objects.filter(
            turno__in=turnos,
            inscrito=True
        ).count()
    else:
        cantidad_inscritos = devotos.count()

    # Nombre de procesión
    procesion_nombre = 'CONSULTA GENERAL'
    if filtro_procesion:
        procesion = Procesion.objects.filter(id=filtro_procesion, activo=True).first()
        if procesion:
            procesion_nombre = procesion.nombre

    # Número de turno
    numero_turno = 'CONSULTA GENERAL'
    if filtro_turno:
        turno = Turno.objects.filter(id=filtro_turno).first()
        if turno:
            numero_turno = str(turno.numero_turno)

    context = {
        'fecha_hoy': fecha_hoy,
        'inscripciones': devotos,
        'cantidad_inscritos': cantidad_inscritos,
        'procesion_nombre': procesion_nombre,
        'numero_turno': numero_turno,
    }

    html = render_to_string('gestion_turnos/reporte_inscripciones_pdf.html', context)
    pdf = HTML(string=html).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_inscripciones.pdf"'
    return response



@login_required
def exportar_inscripciones_excel(request):
    filtro_procesion = request.GET.get('procesion', '')
    filtro_turno = request.GET.get('turno')
    fecha_inicio_str = request.GET.get('fechainicio', '')
    fecha_fin_str = request.GET.get('fechafin', '')

    inscripciones = RegistroInscripcion.objects.filter(inscrito=True)

    if fecha_inicio_str:
        inscripciones = inscripciones.filter(fecha_inscripcion__date__gte=fecha_inicio_str)
    if fecha_fin_str:
        inscripciones = inscripciones.filter(fecha_inscripcion__date__lte=fecha_fin_str)
    if filtro_procesion:
        inscripciones = inscripciones.filter(turno__procesion_id=filtro_procesion)
    if filtro_turno:
        inscripciones = inscripciones.filter(turno_id=filtro_turno)

    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Inscripciones"

    # Encabezados
    encabezados = ["N° Turno", "Devoto", "Fecha Inscripción", "Entregado", "Pagado", "Cambio"]
    ws.append(encabezados)

    # Datos
    for ins in inscripciones.select_related('turno', 'devoto').order_by('turno__numero_turno', 'devoto__nombre'):
        ws.append([
            ins.turno.numero_turno,
            ins.devoto.nombre,
            ins.fecha_inscripcion.strftime('%Y-%m-%d %H:%M:%S'),
            "Sí" if ins.entregado else "No",
            float(ins.monto_pagado),
            float(ins.cambio),
        ])

    # Ajustar ancho de columnas
    for col_idx, header in enumerate(encabezados, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(15, len(header) + 2)

    # Respuesta HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_inscripciones.xlsx"'
    return response


@login_required
def reenviar_correo_inscripcion(request, inscripcion_id):
    inscripcion = get_object_or_404(RegistroInscripcion, id=inscripcion_id)

    try:
        enviado = enviar_confirmacion_inscripcion_correo(
            inscripcion,
            usuario=request.user
        )

        if enviado:
            return JsonResponse({
                "ok": True,
                "message": "📧 Correo reenviado correctamente."
            })
        else:
            return JsonResponse({
                "ok": False,
                "message": "No se encontró un correo válido para enviar."
            })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "message": "Ocurrió un error al reenviar el correo.",
            "error": str(e)
        }, status=500)